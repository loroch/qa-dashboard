"""
Export service: generates CSV and Excel exports from dashboard data.
"""
import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


def flatten_issue(issue: dict) -> dict:
    """Flatten nested Jira issue dict to a flat row for export."""
    assignee = issue.get("assignee") or {}
    qa_owner = issue.get("qa_owner") or {}
    sprint = issue.get("sprint") or {}
    versions = ", ".join(v["name"] for v in issue.get("fix_versions", []))
    components = ", ".join(issue.get("components", []))
    labels = ", ".join(issue.get("labels", []))

    return {
        "Key": issue.get("key", ""),
        "Summary": issue.get("summary", ""),
        "Status": issue.get("status", ""),
        "Priority": issue.get("priority", ""),
        "Issue Type": issue.get("issue_type", ""),
        "Assignee": assignee.get("display_name", ""),
        "QA Owner": qa_owner.get("display_name", ""),
        "Reporter": (issue.get("reporter") or {}).get("display_name", ""),
        "Fix Versions": versions,
        "Components": components,
        "Labels": labels,
        "Epic": issue.get("epic_name") or issue.get("epic_link", ""),
        "Bundle": issue.get("bundle", ""),
        "Activity": issue.get("activity", ""),
        "Sprint": sprint.get("name", ""),
        "Story Points": issue.get("story_points", ""),
        "Test Count": issue.get("test_count", ""),
        "Days in Status": issue.get("days_in_status", 0),
        "Aging Level": issue.get("aging_level", ""),
        "Created": issue.get("created", ""),
        "Updated": issue.get("updated", ""),
        "Due Date": issue.get("due_date", ""),
        "URL": issue.get("url", ""),
    }


def issues_to_csv(issues: list[dict]) -> bytes:
    """Convert list of issues to CSV bytes."""
    if not issues:
        return b""
    rows = [flatten_issue(i) for i in issues]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def issues_to_excel(issues: list[dict], sheet_name: str = "Issues") -> bytes:
    """Convert list of issues to Excel bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed")
        raise

    rows = [flatten_issue(i) for i in issues]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if not rows:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    headers = list(rows[0].keys())

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows with alternating colors
    fill_even = PatternFill("solid", fgColor="EBF0F8")
    aging_fills = {
        "warning": PatternFill("solid", fgColor="FFF3CD"),
        "critical": PatternFill("solid", fgColor="FFD0B5"),
        "overdue": PatternFill("solid", fgColor="F8D7DA"),
    }

    for row_idx, row in enumerate(rows, 2):
        aging = row.get("Aging Level", "ok")
        row_fill = aging_fills.get(aging, fill_even if row_idx % 2 == 0 else None)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            if row_fill:
                cell.fill = row_fill
            # Make URL a hyperlink
            if header == "URL" and row[header]:
                cell.hyperlink = row[header]
                cell.font = Font(color="0563C1", underline="single")

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(rows[r][header])) for r in range(len(rows))), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def changelog_to_excel(entries: list[dict]) -> bytes:
    """Export changelog entries to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Changelog"

    headers = ["ID", "Type", "Component", "Description", "Old Value", "New Value",
               "Changed By", "Changed At", "Version"]
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.get("id"))
        ws.cell(row=row_idx, column=2, value=entry.get("change_type"))
        ws.cell(row=row_idx, column=3, value=entry.get("component"))
        ws.cell(row=row_idx, column=4, value=entry.get("description"))
        ws.cell(row=row_idx, column=5, value=str(entry.get("old_value", "")))
        ws.cell(row=row_idx, column=6, value=str(entry.get("new_value", "")))
        ws.cell(row=row_idx, column=7, value=entry.get("changed_by"))
        ws.cell(row=row_idx, column=8, value=entry.get("changed_at"))
        ws.cell(row=row_idx, column=9, value=entry.get("version"))

    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
