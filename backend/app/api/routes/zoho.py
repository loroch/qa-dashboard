"""
Zoho Desk API routes.
"""
import logging
from typing import Optional, List
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel

from app.services.zoho_service import get_zoho_service
from app.services.zoho_jira_service import get_zoho_jira_service
from app.services.create_bug_service import get_create_bug_service
from app.zoho.client import get_zoho_client

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/zoho", tags=["zoho"])


@router.get("/status")
async def zoho_connection_status():
    """Test Zoho Desk connection."""
    client = get_zoho_client()
    return await client.test_connection()


@router.get("/dashboard")
async def get_zoho_dashboard(
    refresh: bool = Query(False, description="Force cache refresh"),
):
    """Full Zoho Desk dashboard: tickets by department, status, assignee, trends."""
    try:
        svc = get_zoho_service()
        return await svc.get_dashboard_data(force_refresh=refresh)
    except Exception as e:
        logger.error(f"Zoho dashboard error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tickets")
async def get_zoho_tickets(
    refresh: bool = Query(False),
    status: Optional[str] = Query(None, description="Filter by status e.g. Open"),
):
    """List all Zoho Desk tickets."""
    try:
        svc = get_zoho_service()
        if status:
            from app.zoho.mapper import map_tickets
            raw = await svc.client.get_all_tickets(status=status)
            from app.config import get_settings
            return map_tickets(raw, get_settings().zoho_desk_base_url)
        return await svc.get_all_tickets(force_refresh=refresh)
    except Exception as e:
        logger.error(f"Zoho tickets error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/debug/fields")
async def debug_fields():
    """Return all custom field definitions to find API names."""
    try:
        client = get_zoho_client()
        data = await client.get("/api/v1/fields", params={"module": "tickets"}, include_org=True)
        return [
            {
                "label": f.get("displayLabel") or f.get("label"),
                "api_name": f.get("apiName"),
                "type": f.get("type"),
            }
            for f in data.get("data", [])
            if "project" in (f.get("displayLabel") or f.get("label") or "").lower()
            or "bug" in (f.get("displayLabel") or f.get("label") or "").lower()
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/debug/list-sample")
async def debug_list_sample():
    """Return raw cf fields from the first 3 tickets in the list."""
    try:
        client = get_zoho_client()
        tickets = await client.get_tickets(limit=3)
        return [
            {
                "id": t.get("id"),
                "ticketNumber": t.get("ticketNumber"),
                "cf": t.get("cf"),
                "customFields": t.get("customFields"),
            }
            for t in tickets
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/debug/ticket/{ticket_id}")
async def debug_ticket(ticket_id: str):
    """Return raw ticket data to inspect field names."""
    try:
        client = get_zoho_client()
        data = await client.get(f"/api/v1/tickets/{ticket_id}", include_org=True)
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/departments")
async def get_zoho_departments():
    """List Zoho Desk departments (projects)."""
    try:
        svc = get_zoho_service()
        return await svc.get_departments()
    except Exception as e:
        logger.error(f"Zoho departments error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/refresh")
async def refresh_zoho():
    """Force refresh Zoho Desk cache."""
    try:
        svc = get_zoho_service()
        await svc.refresh()
        return {"status": "ok", "message": "Zoho Desk data refreshed"}
    except Exception as e:
        logger.error(f"Zoho refresh error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/by-project")
async def get_project_status_report(refresh: bool = Query(False)):
    """
    Ticket count grouped by Project Name + Status.
    Returns both flat rows and a grouped-by-project summary.
    """
    try:
        svc = get_zoho_jira_service()
        return await svc.get_project_status_report(force_refresh=refresh)
    except Exception as e:
        logger.error(f"Zoho project report error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/linked")
async def get_linked_report(refresh: bool = Query(False)):
    """
    Zoho tickets with Bug ID cross-referenced with Jira:
    Zoho ticket # | Zoho status | Project | Jira status | Fix version | Parent
    """
    try:
        svc = get_zoho_jira_service()
        return await svc.get_linked_report(force_refresh=refresh)
    except Exception as e:
        logger.error(f"Zoho-Jira linked report error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/linked/export/csv")
async def export_linked_report_csv(refresh: bool = Query(False)):
    """Export Zoho-Jira linked report as CSV."""
    try:
        import csv, io
        svc = get_zoho_jira_service()
        rows = await svc.get_linked_report(force_refresh=refresh)
        if not rows:
            return Response(content=b"", media_type="text/csv")
        fields = [
            "zoho_ticket_number", "zoho_subject", "zoho_status", "zoho_project_name",
            "zoho_assignee", "zoho_contact", "zoho_days_open", "zoho_aging_level",
            "bug_id", "jira_key", "jira_status", "jira_fix_versions",
            "jira_parent_key", "jira_parent_summary", "jira_epic",
            "zoho_url", "jira_url",
        ]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            row["jira_fix_versions"] = ", ".join(row.get("jira_fix_versions") or [])
            writer.writerow(row)
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=zoho_jira_linked.csv"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/linked/export/excel")
async def export_linked_report_excel(refresh: bool = Query(False)):
    """Export Zoho-Jira linked report as Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        svc = get_zoho_jira_service()
        rows = await svc.get_linked_report(force_refresh=refresh)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zoho-Jira Linked"

        headers = [
            "Zoho Ticket", "Subject", "Zoho Status", "Project Name",
            "Assignee", "Contact", "Days Open", "Aging",
            "Bug ID", "Jira Key", "Jira Status", "Fix Versions",
            "Parent Key", "Parent Summary", "Epic",
            "Zoho URL", "Jira URL",
        ]
        fields = [
            "zoho_ticket_number", "zoho_subject", "zoho_status", "zoho_project_name",
            "zoho_assignee", "zoho_contact", "zoho_days_open", "zoho_aging_level",
            "bug_id", "jira_key", "jira_status", "jira_fix_versions",
            "jira_parent_key", "jira_parent_summary", "jira_epic",
            "zoho_url", "jira_url",
        ]

        header_fill = PatternFill("solid", fgColor="1F3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        aging_fills = {
            "overdue":  PatternFill("solid", fgColor="F8D7DA"),
            "critical": PatternFill("solid", fgColor="FFD0B5"),
            "warning":  PatternFill("solid", fgColor="FFF3CD"),
        }

        for row_idx, row in enumerate(rows, 2):
            aging = row.get("zoho_aging_level", "ok")
            row_fill = aging_fills.get(aging)
            for col_idx, field in enumerate(fields, 1):
                val = row.get(field, "")
                if isinstance(val, list):
                    val = ", ".join(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if row_fill:
                    cell.fill = row_fill
                if field in ("zoho_url", "jira_url") and val:
                    cell.hyperlink = val
                    cell.font = Font(color="0563C1", underline="single")

        ws.freeze_panes = "A2"
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        output = io.BytesIO()
        wb.save(output)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=zoho_jira_linked.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/create-bug/meta")
async def get_create_bug_meta():
    """
    Return dropdown data for the Create Bug form:
    fix_versions, epics, sprints, priorities.
    """
    try:
        svc = get_create_bug_service()
        return await svc.get_meta()
    except Exception as e:
        logger.error(f"Create-bug meta error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ticket/{ticket_id}/detail")
async def get_zoho_ticket_detail(ticket_id: str):
    """Full Zoho ticket detail with description and attachment list."""
    try:
        svc = get_create_bug_service()
        return await svc.get_zoho_ticket_detail(ticket_id)
    except Exception as e:
        logger.error(f"Zoho ticket detail error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


class CreateBugRequest(BaseModel):
    zoho_ticket_id: str
    zoho_ticket_url: str
    zoho_ticket_number: str = ""
    summary: str
    description: str = ""
    steps_to_reproduce: str = ""
    actual_result: str = ""
    expected_result: str = ""
    severity: str = "Medium"
    environments: List[str] = []
    found_in_version_id: Optional[str] = None
    epic_key: Optional[str] = None
    fix_version_id: Optional[str] = None
    fix_version_name: Optional[str] = None
    priority_name: Optional[str] = None
    sprint_id: Optional[int] = None
    attachment_ids: List[str] = []


@router.post("/create-bug")
async def create_jira_bug(body: CreateBugRequest):
    """Create a Jira bug from a Zoho Desk ticket."""
    try:
        svc = get_create_bug_service()
        result = await svc.create_jira_bug(
            zoho_ticket_id=body.zoho_ticket_id,
            zoho_ticket_url=body.zoho_ticket_url,
            zoho_ticket_number=body.zoho_ticket_number,
            summary=body.summary,
            description=body.description,
            steps_to_reproduce=body.steps_to_reproduce,
            actual_result=body.actual_result,
            expected_result=body.expected_result,
            severity=body.severity,
            environments=body.environments,
            found_in_version_id=body.found_in_version_id,
            epic_key=body.epic_key,
            fix_version_id=body.fix_version_id,
            fix_version_name=body.fix_version_name,
            priority_name=body.priority_name,
            sprint_id=body.sprint_id,
            attachment_ids=body.attachment_ids,
        )
        return result
    except Exception as e:
        logger.error(f"Create-bug error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tickets/export/csv")
async def export_tickets_csv():
    """Export all Zoho tickets as CSV."""
    try:
        import csv, io
        svc = get_zoho_service()
        tickets = await svc.get_all_tickets()
        if not tickets:
            return Response(content=b"", media_type="text/csv")
        fields = ["ticket_number", "subject", "status", "priority", "assignee_name",
                  "department_name", "contact_name", "days_open", "aging_level",
                  "created", "modified", "due_date", "url"]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(tickets)
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=zoho_tickets.csv"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
